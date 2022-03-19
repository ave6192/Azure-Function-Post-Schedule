import datetime
import logging
from operator import concat, le
import tempfile
import os
from io import BytesIO
import azure.functions as func
import requests
from openpyxl import Workbook, load_workbook
from azure.storage.blob import BlobServiceClient, ContentSettings

username = "mm_es_masteruser"
password = "PrquqFN<b3aSw4h"



def main(mytimer: func.TimerRequest) -> None:
    utc_timestamp = datetime.datetime.utcnow().replace(
        tzinfo=datetime.timezone.utc).isoformat()

    if mytimer.past_due:
        logging.info('The timer is past due!')

    
    myExcel, blob_client = getExcelFromSA()

    entries = getAllDigitalEntries()

    
    upload(entries, myExcel, blob_client)

    

    logging.info('Python timer trigger function ran at %s', utc_timestamp)


def getAllDigitalEntries():
    _query = "Digital" 
    first_res = request(0, 20, _query)
    second_res = request(20, first_res[0]-20, _query)

    
    allEntries = concat(first_res[1], second_res[1])
    # allEntries = concat(first_res[1], [])
    return allEntries

def getExcelFromSA():
    STORAGEACCOUNTURL = os.getenv("STORAGEACCOUNTURL")
    STORAGEACCOUNTKEY = os.getenv("STORAGEACCOUNTKEY")
    CONTAINERNAME = os.getenv("CONTAINERNAME")
    BLOBNAME = os.getenv("EXCELNAME")


    blob_service_client_instance = BlobServiceClient(
    account_url=STORAGEACCOUNTURL, credential=STORAGEACCOUNTKEY)

    blob_client_instance = blob_service_client_instance.get_blob_client(
        CONTAINERNAME, BLOBNAME, snapshot=None)
    
    if(blob_client_instance.exists()):
        blob_data = blob_client_instance.download_blob()
        data = blob_data.readall()
        return data, blob_client_instance
    return None, blob_client_instance


def request(_from, _size, _query):
    url = "https://search-mm-dev-ao4q6qdz7r4ovijbe5mlqqdjii.eu-west-2.es.amazonaws.com/liveindexedcontentpageversionviewmodel/_search"
    payload = {"query":{"bool":{"must":[{"term":{"cultureId":127}},{"term":{"pageType":"7"}},{"multi_match":{"query":_query,"type":"phrase","fields":["pageText","keyWords","summaryText","jobSector","title","location","jobRef"]}}]}},"size":_size,"from":_from,"track_scores":"true","sort":["_score"]}

    response = requests.get(url, auth = (username, password), json=payload)
    json = response.json()

    total = getTotal(json)

    logging.info('Response status code is: %s', response.status_code)
    return (total, getEntries(json))



def getTotal(json):
    return json["hits"]["total"]["value"]

def getEntries(json):
    return json["hits"]["hits"]


def upload(entries, myExcel, blob_client):
    file_name = os.getenv("EXCELNAME")
    workbook = None
    if(myExcel is None): #if file doesn't exist
        workbook = createWorkSheet()
    else:
        workbook = load_workbook(filename=BytesIO(myExcel))
    workbook = appendWorkSheet(entries, workbook)
    
    full_file_name = tempfile.gettempdir() + "/" +file_name
    workbook.save(filename= full_file_name)

    with open(full_file_name, "rb") as data:
        blob_client.upload_blob(data, content_settings=ContentSettings(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'), overwrite=True)


def createWorkSheet():
    workbook = Workbook()
    sheet = workbook.active
    headers = ["Id", "Name", "Sector", "Location", "Contract Type", "Job Description", "Job Url"]
    sheet.append(headers)

    return workbook   

def appendWorkSheet(entries, workbook):
    
    sheet = workbook.active
    for entry in entries:

        id = entry["_id"]
        title = entry["_source"]["title"]
        sector = entry["_source"]["sector"][0]
        location =  entry["_source"]["location"]
        contract_type =  entry["_source"]["contractType"]
        job_description =  entry["_source"]["pageText"]
        job_url = entry["_source"]["url"]

        new_row = [id, title, sector, location, contract_type, job_description, job_url]

        if(not checkExists(entry, sheet)):
            sheet.append(new_row)
        
    return workbook


def checkExists(entry, sheet):
    id = entry["_id"]
    id_cells = list(sheet["A"])

    for cell in id_cells:
        if(cell.value == id):
            return True
    return False




